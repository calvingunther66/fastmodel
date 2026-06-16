"""Generate a synthetic schedule workbook to exercise both extraction paths.

Creates ``samples/sample_schedule.xlsx`` with:
  * a "June" sheet — a normal people x dates grid of shift codes (cell path), and
  * a "Scanned" sheet — a rendered schedule pasted as a PNG image (OCR path).

Run:  python tools/make_sample.py
"""

from __future__ import annotations

import datetime as dt
import os

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont

SAMPLES_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "samples")


def _font(size: int):
    try:
        return ImageFont.load_default(size=size)
    except TypeError:  # Pillow < 10 ignores the size argument
        return ImageFont.load_default()


def _make_schedule_png(path: str) -> str:
    width, height = 620, 260
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    title_font = _font(28)
    body_font = _font(24)

    draw.text((20, 15), "Scanned Roster", fill="black", font=title_font)
    rows = [
        "Name     Mon   Tue   Wed",
        "Alice    D     N     OFF",
        "Carl     OFF   D     D",
    ]
    y = 70
    for line in rows:
        draw.text((20, y), line, fill="black", font=body_font)
        y += 55
    image.save(path)
    return path


def build(out_path: str | None = None) -> str:
    os.makedirs(SAMPLES_DIR, exist_ok=True)
    out_path = out_path or os.path.join(SAMPLES_DIR, "sample_schedule.xlsx")

    wb = openpyxl.Workbook()

    # --- Cell-based sheet ---------------------------------------------------
    ws = wb.active
    ws.title = "June"
    ws["A1"] = "Team Schedule - June 2026"
    ws.cell(row=2, column=1, value="Name")
    for col, day in enumerate(range(1, 8), start=2):
        ws.cell(row=2, column=col, value=dt.date(2026, 6, day))

    data = [
        ("John Doe", ["D", "D", "OFF", "N", "N", "OFF", "D"]),
        ("Jane Smith", ["N", "OFF", "D", "D", "D", "OFF", "N"]),
        ("Bob Lee", ["OFF", "D", "D", "N", "OFF", "D", "D"]),
    ]
    for r, (name, codes) in enumerate(data, start=3):
        ws.cell(row=r, column=1, value=name)
        for c, code in enumerate(codes, start=2):
            ws.cell(row=r, column=c, value=code)

    # --- Image-based sheet --------------------------------------------------
    ws_img = wb.create_sheet("Scanned")
    png_path = _make_schedule_png(os.path.join(SAMPLES_DIR, "_scanned_roster.png"))
    ws_img.add_image(XLImage(png_path), "A1")

    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = build()
    print(f"Wrote sample workbook: {path}")
