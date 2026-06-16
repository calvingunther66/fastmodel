"""Extract a schedule grid from a worksheet's cells.

The layout of real-world rosters varies, so this uses light heuristics:

* The **header row** is the row containing the most "date-like" cells
  (real dates, day-of-week names, or numbers 1-31). Its date-like columns
  define the date axis (one column per day).
* The **name column** is the left-most column (before the first date column)
  that holds text labels on the data rows below the header.

Both can be overridden from the CLI (``--header-row`` / ``--name-col``) for
files with an unusual layout. Codes are kept verbatim — nothing is interpreted.
"""

from __future__ import annotations

import datetime as dt

from openpyxl.utils import column_index_from_string, get_column_letter

from .models import Person, SheetResult, Shift

_DAY_PREFIXES = {"mon", "tue", "wed", "thu", "fri", "sat", "sun"}


def _is_date_like(value) -> bool:
    """True if a cell value plausibly labels a calendar day."""
    if isinstance(value, bool):
        return False
    if isinstance(value, (dt.datetime, dt.date)):
        return True
    if isinstance(value, (int, float)):
        return float(value).is_integer() and 1 <= int(value) <= 31
    if isinstance(value, str):
        s = value.strip().lower()
        if not s:
            return False
        if s[:3] in _DAY_PREFIXES:
            return True
        try:
            return 1 <= int(s) <= 31
        except ValueError:
            return False
    return False


def _date_key(value) -> str:
    """Normalize a header cell into a stable key (ISO date when possible)."""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _resolve_name_col(name_col) -> int | None:
    """Convert a user-supplied name-column (letter or 1-based int) to a 0-based index."""
    if name_col is None:
        return None
    if isinstance(name_col, int):
        return name_col - 1
    s = str(name_col).strip()
    if s.isdigit():
        return int(s) - 1
    return column_index_from_string(s.upper()) - 1


def _read_matrix(ws) -> list[list]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _detect_header(matrix: list[list], forced: int | None = None):
    """Return (header_row_index, [date_col_indices]).

    Returns (None, []) when no plausible header is found in auto mode.
    """
    if forced is not None:
        if forced < 0 or forced >= len(matrix):
            return None, []
        cols = [c for c, v in enumerate(matrix[forced]) if _is_date_like(v)]
        if not cols:  # forced row has no date cells: treat every populated column as a day
            cols = [c for c, v in enumerate(matrix[forced]) if v not in (None, "")]
        return forced, cols

    best_idx, best_cols, best_count = None, [], 0
    for i, row in enumerate(matrix):
        cols = [c for c, v in enumerate(row) if _is_date_like(v)]
        if len(cols) > best_count:
            best_idx, best_cols, best_count = i, cols, len(cols)
    if best_count >= 2:
        return best_idx, best_cols
    return None, []


def _detect_name_col(matrix: list[list], header_idx: int, first_date_col: int) -> int:
    """Pick the column (left of the dates) with the most text labels in data rows."""
    if first_date_col <= 0:
        return 0
    best_c, best_count = 0, -1
    for c in range(first_date_col):
        count = sum(
            1
            for r in range(header_idx + 1, len(matrix))
            if c < len(matrix[r])
            and isinstance(matrix[r][c], str)
            and matrix[r][c].strip()
        )
        if count > best_count:
            best_c, best_count = c, count
    return best_c


def has_cell_grid(ws) -> bool:
    """Cheap check used for sheet routing: is there a detectable date header?"""
    header_idx, _ = _detect_header(_read_matrix(ws))
    return header_idx is not None


def extract_cells(ws, header_row: int | None = None, name_col=None) -> SheetResult:
    """Extract people and their per-date raw codes from a worksheet's cells."""
    res = SheetResult(name=ws.title, source_type="cells")
    matrix = _read_matrix(ws)

    forced_header = (header_row - 1) if header_row else None
    header_idx, date_cols = _detect_header(matrix, forced=forced_header)
    if header_idx is None or not date_cols:
        res.source_type = "empty"
        res.warnings.append("no date-like header row detected; no shifts extracted")
        return res

    res.warnings.append(
        f"header row {'forced' if forced_header is not None else 'auto-detected'} "
        f"at row {header_idx + 1}"
    )

    first_date_col = min(date_cols)
    resolved = _resolve_name_col(name_col)
    n_col = resolved if resolved is not None else _detect_name_col(
        matrix, header_idx, first_date_col
    )
    res.warnings.append(f"name column = {get_column_letter(n_col + 1)}")

    header = matrix[header_idx]
    date_keys = [(c, _date_key(header[c])) for c in date_cols]
    res.date_axis = [key for _, key in date_keys]

    for r in range(header_idx + 1, len(matrix)):
        row = matrix[r]
        name = row[n_col] if n_col < len(row) else None
        if name is None or not str(name).strip():
            continue
        person = Person(name=str(name).strip())
        for c, key in date_keys:
            val = row[c] if c < len(row) else None
            if val is None or not str(val).strip():
                continue
            person.shifts.append(Shift(date=key, raw_code=str(val).strip()))
        if person.shifts:
            res.people.append(person)

    if not res.people:
        res.warnings.append("header detected but no person rows with codes were found")
    return res
