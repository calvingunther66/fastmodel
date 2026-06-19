"""Persistence for the uploaded schedule and per-person calendar tokens.

State lives in DATA_DIR:
  current.xlsx   the most recently uploaded workbook (raw bytes)
  schedule.json  the parsed result + metadata (which sheet, when, sheet list)
  tokens.json    stable {person_name: secret_token} used for live .ics URLs

Tokens are kept stable across re-uploads so a person's subscribed calendar link
never breaks when a new schedule is posted.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import secrets
import threading
from pathlib import Path

import openpyxl

from schedule_extractor.definitions import decode, shift_window
from schedule_extractor.roster_extractor import extract_roster

from .config import DATA_DIR
from .coverage import apply_overrides, apply_swaps, find_open_shift

_LEVEL_TO_TYPE = {"day": "day", "mid": "midshift", "night": "night"}

# A sheet looks like a draft/working copy if it starts with a code like "KH-2",
# "NEW-3", "OLD-3" or ends with a copy marker like "(2)".
_DRAFT_RE = re.compile(r"^[A-Za-z]{1,5}-?\d|\(\d+\)\s*$")


def is_draft_sheet(name: str) -> bool:
    return bool(_DRAFT_RE.search(name or ""))


class ScheduleStore:
    def __init__(self, data_dir: Path = DATA_DIR) -> None:
        self.data_dir = data_dir
        self.xlsx_path = data_dir / "current.xlsx"
        self.schedule_path = data_dir / "schedule.json"
        self.tokens_path = data_dir / "tokens.json"
        self.overrides_path = data_dir / "overrides.json"
        self.contacts_path = data_dir / "contacts.json"
        self.offers_path = data_dir / "availability.json"
        self.stats_path = data_dir / "stats.json"
        self.settings_path = data_dir / "settings.json"
        self.prefs_path = data_dir / "prefs.json"
        self.swaps_path = data_dir / "swaps.json"
        self.templates_path = data_dir / "templates.json"
        self.vacations_path = data_dir / "vacations.json"
        self.holidays_path = data_dir / "holidays.json"
        self._lock = threading.Lock()

    # ---- low-level json helpers ------------------------------------------
    def _read_json(self, path: Path, default):
        if path.exists():
            return json.loads(path.read_text())
        return default

    def _write_json(self, path: Path, data) -> None:
        path.write_text(json.dumps(data, indent=2, default=str))

    # ---- parsing ----------------------------------------------------------
    def _parse(self, sheet_name: str | None):
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        sheets = wb.sheetnames
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Auto-pick by (named people, total shifts, not-a-draft); on a tie
            # prefer the right-most sheet (final versions tend to sit last) and
            # de-prioritise obvious draft tabs (e.g. "KH-2…", "…(2)").
            best_score, ws = (-1, -1, -1, -1), wb.worksheets[0]
            for index, cand in enumerate(wb.worksheets):
                try:
                    parsed = extract_roster(cand)
                    named = sum(1 for p in parsed["people"] if p["name"])
                    total = sum(len(p["shifts"]) for p in parsed["people"])
                except Exception:
                    named = total = -1
                score = (named, total, 0 if is_draft_sheet(cand.title) else 1, index)
                if score > best_score:
                    best_score, ws = score, cand
        result = extract_roster(ws)
        result["available_sheets"] = sheets
        result["parsed_sheet"] = ws.title
        result["uploaded_at"] = dt.datetime.now(dt.timezone.utc).isoformat()
        return result

    # ---- public API -------------------------------------------------------
    def ingest(self, xlsx_bytes: bytes, sheet_name: str | None = None) -> dict:
        """Store a freshly uploaded workbook and parse it."""
        with self._lock:
            self.xlsx_path.write_bytes(xlsx_bytes)
            result = self._parse(sheet_name)
            self._write_json(self.schedule_path, result)
            self._ensure_tokens(result)
            self._record_work(result)
            return result

    def reparse(self, sheet_name: str) -> dict:
        """Re-parse the already-uploaded workbook with a specific sheet."""
        with self._lock:
            if not self.xlsx_path.exists():
                raise FileNotFoundError("no workbook uploaded yet")
            result = self._parse(sheet_name)
            self._write_json(self.schedule_path, result)
            self._ensure_tokens(result)
            self._record_work(result)
            return result

    def create_schedule(self, title: str, start: str, end: str,
                        assignments: dict, contacts: dict | None = None) -> dict:
        """Build a schedule dict from in-app assignments (no .xlsx) and store it.

        `assignments` is {person: {date: {level: code}}} where level is
        day|mid|night. Codes are decoded to meaning + times via definitions, so a
        created schedule is identical in shape to a parsed one.
        """
        with self._lock:
            people = []
            for name, days in (assignments or {}).items():
                shifts = []
                for date, levels in (days or {}).items():
                    for level, code in (levels or {}).items():
                        code = (code or "").strip()
                        if not code:
                            continue
                        st = _LEVEL_TO_TYPE.get(level, "day")
                        info = decode(code)
                        s_t, e_t, cross = shift_window(code, st)
                        shifts.append({
                            "date": date, "code": code, "shift_type": st,
                            "category": info["category"], "meaning": info["meaning"],
                            "start": s_t, "end": e_t, "crosses_midnight": cross,
                            "available": True,
                        })
                shifts.sort(key=lambda s: (s["date"], s["shift_type"]))
                people.append({
                    "name": name, "contact": (contacts or {}).get(name, []),
                    "shifts": shifts, "unavailable": [], "notes": [],
                })
            result = {
                "sheet": title, "parsed_sheet": title, "available_sheets": [title],
                "date_range": {"start": start, "end": end},
                "uploaded_at": dt.datetime.now(dt.timezone.utc).isoformat(),
                "created": True, "people": people, "warnings": [],
            }
            self._write_json(self.schedule_path, result)
            self._ensure_tokens(result)
            self._record_work(result)
            return result

    def get_raw_schedule(self) -> dict | None:
        """The parsed schedule as stored, without call-out overrides applied."""
        return self._read_json(self.schedule_path, None)

    def get_schedule(self) -> dict | None:
        """The schedule with call-outs flagged, covers injected, contact overrides
        applied. Always a fresh copy (safe to mutate by callers)."""
        base = self._read_json(self.schedule_path, None)
        if base is None:
            return None
        sched = apply_overrides(base, self._callouts())  # deep-copies even if empty
        apply_swaps(sched, self._swaps())                 # approved swaps move shifts
        self._apply_vacation_decisions(sched)             # admin approve/deny over green-fill
        contacts = self._contacts()
        if contacts:
            for p in sched.get("people", []):
                if p.get("name") in contacts:
                    p["contact"] = contacts[p["name"]]
        return sched

    # ---- contact overrides (members edit their own) ----------------------
    def _contacts(self) -> dict:
        return self._read_json(self.contacts_path, {})

    def set_contact(self, person: str, lines: list[str]) -> None:
        with self._lock:
            contacts = self._contacts()
            contacts[person] = [str(line).strip() for line in lines if str(line).strip()]
            self._write_json(self.contacts_path, contacts)

    # ---- availability offers (people offering to cover) ------------------
    def _offers(self) -> list[dict]:
        return self._read_json(self.offers_path, {}).get("offers", [])

    def list_offers(self) -> list[dict]:
        return self._offers()

    def add_offer(self, person: str, date: str, note: str = "") -> None:
        with self._lock:
            offers = self._offers()
            if not any(o["person"] == person and o["date"] == date for o in offers):
                offers.append({"person": person, "date": date, "note": note})
                self._write_json(self.offers_path, {"offers": offers})

    def remove_offer(self, person: str, date: str) -> None:
        with self._lock:
            offers = [o for o in self._offers()
                      if not (o["person"] == person and o["date"] == date)]
            self._write_json(self.offers_path, {"offers": offers})

    def offers_for_date(self, date: str) -> set[str]:
        return {o["person"] for o in self._offers() if o["date"] == date}

    # ---- member availability preferences (B4) ----------------------------
    def _prefs(self) -> dict:
        return self._read_json(self.prefs_path, {})

    def list_prefs(self) -> dict:
        return self._prefs()

    def get_prefs(self, person: str) -> dict:
        return self._prefs().get(person, {})

    def set_prefs(self, person: str, prefs: dict) -> dict:
        clean = {
            "no_weekdays": sorted({int(d) for d in (prefs.get("no_weekdays") or [])
                                   if str(d).isdigit() and 0 <= int(d) <= 6}),
            "prefer_nights": bool(prefs.get("prefer_nights")),
            "max_consecutive": int(prefs.get("max_consecutive") or 0),
            "reminder_minutes": int(prefs.get("reminder_minutes") or 0),
            "note": str(prefs.get("note") or "").strip(),
        }
        with self._lock:
            data = self._prefs()
            data[person] = clean
            self._write_json(self.prefs_path, data)
        return clean

    # ---- open-shift claims (B1) ------------------------------------------
    def _claims(self) -> list[dict]:
        return self._read_json(self.overrides_path, {}).get("claims", [])

    def list_claims(self) -> list[dict]:
        return self._claims()

    def add_claim(self, claimer: str, name: str, date: str, shift_type: str) -> None:
        with self._lock:
            data = self._read_json(self.overrides_path, {})
            claims = data.get("claims", [])
            if not any(c["claimer"] == claimer and self._same(c, name, date, shift_type)
                       for c in claims):
                claims.append({"claimer": claimer, "name": name, "date": date,
                               "shift_type": shift_type})
            data["claims"] = claims
            self._write_json(self.overrides_path, data)

    def remove_claim(self, claimer: str, name: str, date: str, shift_type: str) -> None:
        with self._lock:
            data = self._read_json(self.overrides_path, {})
            data["claims"] = [c for c in data.get("claims", [])
                              if not (c["claimer"] == claimer
                                      and self._same(c, name, date, shift_type))]
            self._write_json(self.overrides_path, data)

    def claims_for(self, name: str, date: str, shift_type: str) -> list[str]:
        return [c["claimer"] for c in self._claims()
                if self._same(c, name, date, shift_type)]

    # ---- shift swaps (B3) -------------------------------------------------
    def _swaps(self) -> list[dict]:
        return self._read_json(self.swaps_path, {}).get("swaps", [])

    def list_swaps(self) -> list[dict]:
        return self._swaps()

    def propose_swap(self, a_person, a_date, a_type, b_person, b_date, b_type) -> dict:
        sw = {
            "id": secrets.token_hex(6), "status": "proposed",
            "a_person": a_person, "a_date": a_date, "a_type": a_type,
            "b_person": b_person, "b_date": b_date, "b_type": b_type,
        }
        with self._lock:
            swaps = self._swaps()
            swaps.append(sw)
            self._write_json(self.swaps_path, {"swaps": swaps})
        return sw

    def set_swap_status(self, swap_id: str, status: str) -> dict | None:
        with self._lock:
            swaps = self._swaps()
            found = None
            for sw in swaps:
                if sw["id"] == swap_id:
                    sw["status"] = status
                    found = sw
                    break
            self._write_json(self.swaps_path, {"swaps": swaps})
            return found

    # ---- vacation approvals (H1) -----------------------------------------
    # The workbook marks approved vacation by a green fill (shift["approved"]).
    # Admins can override per (person, date) here; decisions win over the fill.
    @staticmethod
    def _vac_key(person: str, date: str) -> str:
        return f"{person}|{date}"

    def _vacations(self) -> dict:
        path = getattr(self, "vacations_path", None)
        return self._read_json(path, {}) if path else {}

    @staticmethod
    def _is_vacation(shift: dict) -> bool:
        return (shift.get("code") or "").upper() == "V"

    def _vacation_status(self, shift: dict, decision: str | None) -> str:
        """Effective status: admin decision wins, else the workbook's green fill."""
        if decision in ("approved", "denied"):
            return decision
        return "approved" if shift.get("approved") else "pending"

    def _apply_vacation_decisions(self, sched: dict) -> None:
        """Stamp each V shift with its effective approval (decisions override fill)."""
        decisions = self._vacations()
        for p in sched.get("people", []):
            name = p.get("name")
            for s in p.get("shifts", []):
                if not self._is_vacation(s):
                    continue
                status = self._vacation_status(s, decisions.get(self._vac_key(name, s["date"])))
                s["vacation_status"] = status
                s["approved"] = status == "approved"

    def list_vacations(self) -> list[dict]:
        """All vacation (V) entries in the active schedule with effective status."""
        sched = self.get_raw_schedule() or {}
        decisions = self._vacations()
        out = []
        for p in sched.get("people", []):
            name = p.get("name")
            if not name:
                continue
            for s in p.get("shifts", []):
                if not self._is_vacation(s):
                    continue
                decision = decisions.get(self._vac_key(name, s["date"]))
                out.append({
                    "person": name, "date": s["date"],
                    "from_workbook": bool(s.get("approved")),
                    "decision": decision,
                    "status": self._vacation_status(s, decision),
                })
        out.sort(key=lambda v: (v["date"], v["person"]))
        return out

    def set_vacation(self, person: str, date: str, status: str) -> dict:
        """Record an approve/deny decision, or clear it (status 'pending')."""
        with self._lock:
            data = self._vacations()
            key = self._vac_key(person, date)
            if status == "pending":
                data.pop(key, None)
            elif status in ("approved", "denied"):
                data[key] = status
            else:
                raise ValueError("status must be approved|denied|pending")
            self._write_json(self.vacations_path, data)
        return {"person": person, "date": date, "status": status}

    def approved_vacations(self) -> dict:
        """{NAME_UPPER: {dates}} a person is on *approved* vacation (engine block)."""
        out: dict[str, set] = {}
        for v in self.list_vacations():
            if v["status"] == "approved":
                out.setdefault(v["person"].upper(), set()).add(v["date"])
        return out

    # ---- holiday registry (H3) -------------------------------------------
    # Unit holidays (dates), admin-managed. Drives grid highlighting and the
    # "worked a holiday" equity metric. Distinct from a person's H status code
    # (which means *they* have that holiday off).
    def _holidays(self) -> dict:
        path = getattr(self, "holidays_path", None)
        return self._read_json(path, {}) if path else {}

    def list_holidays(self) -> list[dict]:
        return [{"date": d, "label": lbl}
                for d, lbl in sorted(self._holidays().items())]

    def holiday_dates(self) -> set:
        return set(self._holidays().keys())

    def add_holiday(self, date: str, label: str = "") -> dict:
        try:
            dt.date.fromisoformat(date)
        except (ValueError, TypeError):
            raise ValueError("date must be YYYY-MM-DD")
        with self._lock:
            data = self._holidays()
            data[date] = str(label or "").strip()
            self._write_json(self.holidays_path, data)
        return {"date": date, "label": data[date]}

    def remove_holiday(self, date: str) -> None:
        with self._lock:
            data = self._holidays()
            data.pop(date, None)
            self._write_json(self.holidays_path, data)

    # ---- builder templates (C3) ------------------------------------------
    def list_templates(self) -> dict:
        return self._read_json(self.templates_path, {})

    def save_template(self, name: str, weekday_levels: dict) -> dict:
        with self._lock:
            data = self.list_templates()
            data[name] = weekday_levels
            self._write_json(self.templates_path, data)
        return {name: weekday_levels}

    def delete_template(self, name: str) -> None:
        with self._lock:
            data = self.list_templates()
            data.pop(name, None)
            self._write_json(self.templates_path, data)

    # ---- adaptive stats: work frequency + cover step-ups -----------------
    def _stats(self) -> dict:
        return self._read_json(self.stats_path, {"work_by_period": {}, "covers": {}})

    def _record_work(self, result: dict) -> None:
        """Fold this schedule period's worked-shift counts into the stats store.

        Keyed by period so re-parsing the same period replaces (not double-counts);
        distinct periods accumulate, so the picture sharpens as more are uploaded.
        """
        # Key by the schedule's date range so re-parsing the same period (e.g.
        # switching from a draft sheet to the final one) replaces rather than
        # double-counts; genuinely new periods accumulate.
        dr = result.get("date_range", {}) or {}
        period = f"{dr.get('start')}..{dr.get('end')}"
        counts: dict = {}
        for p in result.get("people", []):
            name = p.get("name")
            if not name:
                continue
            by_code, by_type, total = {}, {}, 0
            for s in p["shifts"]:
                if s.get("category") == "location":
                    code = s["code"].upper()
                    st = s["shift_type"]
                    by_code[code] = by_code.get(code, 0) + 1
                    by_type[st] = by_type.get(st, 0) + 1
                    total += 1
            if total:
                counts[name] = {"by_code": by_code, "by_type": by_type, "total": total}
        stats = self._stats()
        stats["work_by_period"][period] = counts
        self._write_json(self.stats_path, stats)

    def _bump_cover(self, person: str, code: str | None, shift_type: str) -> None:
        stats = self._stats()
        c = stats.setdefault("covers", {}).setdefault(
            person, {"count": 0, "by_code": {}, "by_type": {}})
        c["count"] += 1
        if code:
            c["by_code"][code.upper()] = c["by_code"].get(code.upper(), 0) + 1
        if shift_type:
            c["by_type"][shift_type] = c["by_type"].get(shift_type, 0) + 1
        self._write_json(self.stats_path, stats)

    def _unbump_cover(self, person: str, code: str | None, shift_type: str) -> None:
        """Reverse a cover step-up (for undo), never dropping below zero."""
        stats = self._stats()
        c = stats.get("covers", {}).get(person)
        if not c:
            return
        c["count"] = max(0, c.get("count", 0) - 1)
        if code and code.upper() in c.get("by_code", {}):
            c["by_code"][code.upper()] = max(0, c["by_code"][code.upper()] - 1)
        if shift_type and shift_type in c.get("by_type", {}):
            c["by_type"][shift_type] = max(0, c["by_type"][shift_type] - 1)
        self._write_json(self.stats_path, stats)

    def aggregated_stats(self) -> dict:
        """Per-person totals across all periods, plus cover step-up counts."""
        stats = self._stats()
        work: dict = {}
        for counts in stats.get("work_by_period", {}).values():
            for name, c in counts.items():
                w = work.setdefault(name, {"by_code": {}, "by_type": {}, "total": 0})
                for k, v in c["by_code"].items():
                    w["by_code"][k] = w["by_code"].get(k, 0) + v
                for k, v in c["by_type"].items():
                    w["by_type"][k] = w["by_type"].get(k, 0) + v
                w["total"] += c["total"]
        return {
            "work": work,
            "covers": stats.get("covers", {}),
            "periods": len(stats.get("work_by_period", {})),
        }

    def _equity(self) -> dict:
        """Per-person nights/weekends/holidays/hours from the active schedule (D1).

        History stats keep only counts (no dates), so the weekend/hours breakdown is
        computed from the current period's schedule, which retains dates + times."""
        from .validate import shift_hours  # local import avoids a cycle
        sched = self.get_raw_schedule() or {}
        holiday_dates = self.holiday_dates()
        out: dict[str, dict] = {}
        for p in sched.get("people", []):
            name = p.get("name")
            if not name:
                continue
            nights = weekends = holidays = holidays_worked = 0
            hours = 0.0
            for s in p.get("shifts", []):
                if (s.get("code") or "").upper() == "H":
                    holidays += 1
                if s.get("category") != "location":
                    continue
                hours += shift_hours(s)
                if s.get("shift_type") == "night":
                    nights += 1
                if s.get("date") in holiday_dates:
                    holidays_worked += 1  # worked a registered unit holiday
                try:
                    if dt.date.fromisoformat(s["date"]).weekday() >= 5:
                        weekends += 1
                except (ValueError, KeyError, TypeError):
                    pass
            out[name] = {"nights": nights, "weekends": weekends,
                         "holidays": holidays, "holidays_worked": holidays_worked,
                         "hours": round(hours, 1)}
        return out

    def period_trend(self) -> list[dict]:
        """Total worked shifts per period, oldest first (period keys are
        'start..end' so lexical order is chronological). Powers the trend chart (K1)."""
        out = []
        for period, counts in self._stats().get("work_by_period", {}).items():
            out.append({
                "period": period,
                "shifts": sum(c.get("total", 0) for c in counts.values()),
                "people": len(counts),
            })
        out.sort(key=lambda r: r["period"])
        return out

    def leaderboard(self) -> dict:
        """Per-person insights: cover step-ups plus an equity breakdown (D1)."""
        agg = self.aggregated_stats()
        covers, work = agg["covers"], agg["work"]
        equity = self._equity()
        names = set(covers) | set(work)
        # Include everyone currently on the schedule, even with zero history.
        sched = self.get_raw_schedule() or {}
        names |= {p["name"] for p in sched.get("people", []) if p.get("name")}
        rows = []
        for n in sorted(names):
            c = covers.get(n, {})
            w = work.get(n, {})
            eq = equity.get(n, {})
            rows.append({
                "name": n,
                "covers": c.get("count", 0),
                "covers_by_type": c.get("by_type", {}),
                "covers_by_code": c.get("by_code", {}),
                "worked_total": w.get("total", 0),
                "worked_by_code": w.get("by_code", {}),
                "nights": eq.get("nights", 0),
                "weekends": eq.get("weekends", 0),
                "holidays": eq.get("holidays", 0),
                "holidays_worked": eq.get("holidays_worked", 0),
                "hours": eq.get("hours", 0.0),
            })
        rows.sort(key=lambda r: (-r["covers"], -r["worked_total"], r["name"]))
        return {"periods": agg["periods"], "people": rows, "trend": self.period_trend()}

    # ---- tunable settings -------------------------------------------------
    def _settings(self) -> dict:
        return self._read_json(self.settings_path, {})

    def get_fairness_weight(self) -> float:
        try:
            return max(0.0, min(1.0, float(self._settings().get("fairness_weight", 0.5))))
        except (TypeError, ValueError):
            return 0.5

    def set_fairness_weight(self, value: float) -> float:
        w = max(0.0, min(1.0, float(value)))
        with self._lock:
            s = self._settings()
            s["fairness_weight"] = w
            self._write_json(self.settings_path, s)
        return w

    # ---- call-out overrides ----------------------------------------------
    def _callouts(self) -> list[dict]:
        return self._read_json(self.overrides_path, {}).get("callouts", [])

    def _save_callouts(self, callouts: list[dict]) -> None:
        self._write_json(self.overrides_path, {"callouts": callouts})

    @staticmethod
    def _same(co: dict, name: str, date: str, shift_type: str) -> bool:
        return co["name"] == name and co["date"] == date and co["shift_type"] == shift_type

    def list_callouts(self) -> list[dict]:
        return self._callouts()

    def mark_sick(self, name: str, date: str, shift_type: str, code: str | None = None,
                  reason: str = "out sick") -> None:
        with self._lock:
            callouts = self._callouts()
            if not any(self._same(c, name, date, shift_type) for c in callouts):
                callouts.append({
                    "name": name, "date": date, "shift_type": shift_type,
                    "code": code, "reason": reason, "covered_by": None,
                    "created_at": dt.datetime.now(dt.timezone.utc).isoformat(),
                })
                self._save_callouts(callouts)

    def assign_cover(self, name: str, date: str, shift_type: str, covered_by: str,
                     code: str | None = None) -> None:
        with self._lock:
            callouts = self._callouts()
            for c in callouts:
                if self._same(c, name, date, shift_type):
                    c["covered_by"] = covered_by
                    code = code or c.get("code")
                    break
            else:
                callouts.append({
                    "name": name, "date": date, "shift_type": shift_type,
                    "code": code, "reason": "out sick", "covered_by": covered_by,
                })
            self._save_callouts(callouts)
            # Record the step-up so the scorer learns who reliably covers.
            if code is None:
                shift = find_open_shift(self.get_raw_schedule() or {}, name, date, shift_type)
                code = (shift or {}).get("code")
            self._bump_cover(covered_by, code, shift_type)

    def assign_cascade(self, name: str, date: str, shift_type: str,
                       mover: str, from_code: str | None, from_type: str,
                       backfill: str) -> None:
        """Apply a two-step cascade: mover covers the open shift, a free person
        backfills the mover's vacated slot."""
        # 1) mover covers the open (sick) shift
        self.assign_cover(name, date, shift_type, mover)
        # 2) the mover's own slot becomes a call-out, backfilled by `backfill`
        self.mark_sick(mover, date, from_type, code=from_code,
                       reason=f"moved to cover {name}")
        self.assign_cover(mover, date, from_type, backfill, code=from_code)

    def unassign_cover(self, name: str, date: str, shift_type: str) -> bool:
        """Undo a cover assignment: the shift goes back to open (call-out kept).

        Also reverses the learned step-up so the fairness picture isn't skewed by a
        mistaken assignment. Returns True if a cover was actually removed."""
        with self._lock:
            callouts = self._callouts()
            for c in callouts:
                if self._same(c, name, date, shift_type) and c.get("covered_by"):
                    prev = c["covered_by"]
                    c["covered_by"] = None
                    self._save_callouts(callouts)
                    self._unbump_cover(prev, c.get("code"), shift_type)
                    return True
        return False

    def apply_chain(self, name: str, date: str, shift_type: str,
                    steps: list[dict], backfill: str) -> None:
        """Apply an N-step cascade (I2). steps[i] moves a person onto the slot the
        previous step vacated; the final vacated slot is filled by `backfill`.

        step = {mover, from: {code, shift_type}, onto_code, onto_type}. step[0]
        moves `mover` onto the open (name/date/shift_type) shift."""
        if not steps:
            return
        # 1) first mover covers the open shift
        self.assign_cover(name, date, shift_type, steps[0]["mover"],
                          code=steps[0].get("onto_code"))
        # 2) each mover vacates their slot; the next mover (or backfill) fills it
        for i, st in enumerate(steps):
            frm = st["from"]
            filler = steps[i + 1]["mover"] if i + 1 < len(steps) else backfill
            self.mark_sick(st["mover"], date, frm["shift_type"], code=frm.get("code"),
                           reason=f"moved to cover {st.get('onto_code') or 'shift'}")
            self.assign_cover(st["mover"], date, frm["shift_type"], filler,
                              code=frm.get("code"))

    def clear_callout(self, name: str, date: str, shift_type: str) -> None:
        with self._lock:
            removed = [c for c in self._callouts() if self._same(c, name, date, shift_type)]
            callouts = [c for c in self._callouts()
                        if not self._same(c, name, date, shift_type)]
            self._save_callouts(callouts)
        # If the cleared call-out had a cover, reverse its learned step-up too.
        for c in removed:
            if c.get("covered_by"):
                self._unbump_cover(c["covered_by"], c.get("code"), shift_type)

    def _tokens(self) -> dict:
        return self._read_json(self.tokens_path, {})

    def _ensure_tokens(self, schedule: dict) -> None:
        tokens = self._tokens()
        changed = False
        for person in schedule.get("people", []):
            name = person.get("name")
            if name and name not in tokens:
                tokens[name] = secrets.token_urlsafe(16)
                changed = True
        if changed:
            self._write_json(self.tokens_path, tokens)

    def people_with_tokens(self) -> list[dict]:
        schedule = self.get_schedule() or {}
        tokens = self._tokens()
        out = []
        for person in schedule.get("people", []):
            name = person.get("name")
            if not name:
                continue
            out.append({
                "name": name,
                "token": tokens.get(name),
                "shift_count": len(person.get("shifts", [])),
            })
        return out

    def person_by_token(self, token: str) -> dict | None:
        tokens = self._tokens()
        name = next((n for n, t in tokens.items() if t == token), None)
        if not name:
            return None
        schedule = self.get_schedule() or {}
        for person in schedule.get("people", []):
            if person.get("name") == name:
                return person
        return {"name": name, "shifts": []}  # known person, not in current schedule
