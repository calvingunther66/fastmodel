import openpyxl

from schedule_extractor.roster_extractor import extract_roster


def _roster_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 30 - July 2, 26"  # drives start month June + year 2026
    # Date header (row 2) with a month rollover 30 -> 1.
    ws.cell(2, 1, "DATE")
    ws.cell(2, 2, 30)   # Jun 30
    ws.cell(2, 3, 1)    # Jul 1
    ws.cell(2, 4, 2)    # Jul 2
    ws.cell(3, 1, "DAY")
    # Person block: top row = day, middle = midshift, bottom = night.
    ws.cell(4, 1, "SMITH")
    ws.cell(4, 2, "BC")            # day shift, Jun 30
    ws.cell(4, 5, "covers AM")     # note beyond the date columns
    ws.cell(5, 1, "C: 555-1212")
    ws.cell(5, 2, "T")             # midshift on Jun 30 -> split day (BC then T)
    ws.cell(6, 1, "P: TXT")
    ws.cell(6, 3, "HC")            # night shift, Jul 1
    ws.cell(6, 4, "long free text note here")  # note inside a date col -> note
    return ws


def test_roster_levels_and_enrichment():
    res = extract_roster(_roster_ws(), default_year=2026)
    assert res["date_range"] == {"start": "2026-06-30", "end": "2026-07-02"}
    assert len(res["people"]) == 1
    smith = res["people"][0]
    assert smith["name"] == "SMITH"
    assert smith["contact"] == ["C: 555-1212", "P: TXT"]

    by_key = {(s["date"], s["code"]): s for s in smith["shifts"]}

    # Three vertical levels classify to day / midshift / night.
    assert by_key[("2026-06-30", "BC")]["shift_type"] == "day"
    assert by_key[("2026-06-30", "T")]["shift_type"] == "midshift"
    assert by_key[("2026-07-01", "HC")]["shift_type"] == "night"

    # Day + mid on the same date -> split_day flagged on both.
    assert by_key[("2026-06-30", "BC")]["split_day"] is True
    assert by_key[("2026-06-30", "T")]["split_day"] is True

    # Decoding + timing.
    assert by_key[("2026-06-30", "BC")]["meaning"] == "Birth Center"
    assert by_key[("2026-07-01", "HC")]["meaning"] == "Hillcrest"
    night = by_key[("2026-07-01", "HC")]
    assert (night["start"], night["end"], night["crosses_midnight"]) == (
        "19:30", "08:00", True,
    )
    triage = by_key[("2026-06-30", "T")]
    assert (triage["start"], triage["end"]) == ("07:30", "18:00")

    note_texts = {n["text"] for n in smith["notes"]}
    assert "covers AM" in note_texts                  # right of the calendar
    assert "long free text note here" in note_texts   # non-code in a date cell


def test_location_day_windows():
    """BC/HC days use legend hours; clinics default to a full 8-5 day."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 21 - June 24, 26"
    ws.cell(2, 1, "DATE")
    for col, day in [(2, 21), (3, 22), (4, 23)]:
        ws.cell(2, col, day)
    ws.cell(4, 1, "DOE")
    ws.cell(4, 2, "BC")   # Birth Center day
    ws.cell(4, 3, "HC")   # Hillcrest day
    ws.cell(4, 4, "CV")   # Convoy = clinic -> full day

    by = {s["code"]: s for s in extract_roster(ws)["people"][0]["shifts"]}
    assert (by["BC"]["start"], by["BC"]["end"]) == ("07:30", "20:00")
    assert (by["HC"]["start"], by["HC"]["end"]) == ("07:00", "19:30")
    assert (by["CV"]["start"], by["CV"]["end"]) == ("08:00", "17:00")
    assert by["CV"]["meaning"] == "Convoy"


def test_clinic_split_morning_afternoon():
    """A coloured center bar splits a clinic day into morning + afternoon halves."""
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 21 - June 22, 26"
    ws.cell(2, 1, "DATE")
    ws.cell(2, 2, 21)
    ws.cell(4, 1, "DOE")
    ws.cell(4, 2, "MOS")                      # clinic, day row -> morning when split
    ws.cell(5, 1, "C: x")
    mid = ws.cell(5, 2, "CV")                 # mid row -> afternoon
    mid.fill = PatternFill("solid", fgColor="FFFFFF00")  # coloured center bar
    ws.cell(6, 1, "P: y")

    by = {(s["code"], s["shift_type"]): s for s in extract_roster(ws)["people"][0]["shifts"]}
    morning = by[("MOS", "day")]
    afternoon = by[("CV", "midshift")]
    assert (morning["start"], morning["end"]) == ("08:00", "12:00")
    assert morning["split_day"] is True
    assert (afternoon["start"], afternoon["end"]) == ("13:00", "17:00")


def test_ok_is_alias_for_available():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 21 - June 22, 26"
    ws.cell(2, 1, "DATE")
    ws.cell(2, 2, 21)
    ws.cell(4, 1, "DOE")
    ws.cell(4, 2, "OK")
    shift = extract_roster(ws)["people"][0]["shifts"][0]
    assert shift["category"] == "status"
    assert shift["meaning"] == "Available / on-call pool"
    assert shift["start"] is None  # availability marker, no clock window


def test_vacation_approved_by_green_fill():
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 21 - June 22, 26"
    ws.cell(2, 1, "DATE")
    ws.cell(2, 2, 21)
    ws.cell(2, 3, 22)
    ws.cell(4, 1, "JONES")
    approved = ws.cell(4, 2, "V")
    approved.fill = PatternFill("solid", fgColor="FF00B050")  # green -> approved
    ws.cell(4, 3, "V")  # no fill -> not approved

    res = extract_roster(ws, default_year=2026)
    by_date = {s["date"]: s for s in res["people"][0]["shifts"]}
    assert by_date["2026-06-21"]["approved"] is True
    assert by_date["2026-06-22"]["approved"] is False


def test_roster_no_marks_unavailable():
    """A 'no' on a date flags the person unavailable and the shift uncoverable."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 24 - June 25, 26"
    ws.cell(2, 1, "DATE")
    ws.cell(2, 2, 24)
    ws.cell(2, 3, 25)
    ws.cell(4, 1, "CORTES")
    ws.cell(4, 2, "no")     # day row, Jun 24 -> unavailable flag
    ws.cell(4, 3, "R")      # day row, Jun 25 -> normal shift
    ws.cell(5, 1, "C: x")
    ws.cell(6, 1, "P: y")
    ws.cell(6, 2, "BC")     # night row, Jun 24 -> the shift that needs covering

    res = extract_roster(ws, default_year=2026)
    cortes = res["people"][0]
    assert cortes["unavailable"] == [
        {"date": "2026-06-24", "reason": "not available / out sick"}
    ]
    by_key = {(s["date"], s["code"]): s for s in cortes["shifts"]}
    assert by_key[("2026-06-24", "BC")]["available"] is False  # needs coverage
    assert by_key[("2026-06-24", "BC")]["shift_type"] == "night"
    assert by_key[("2026-06-25", "R")]["available"] is True
    # 'no' itself is never emitted as a shift code
    assert all(s["code"].lower() != "no" for s in cortes["shifts"])
