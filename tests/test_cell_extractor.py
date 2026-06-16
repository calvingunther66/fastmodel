import datetime as dt

import openpyxl

from schedule_extractor.cell_extractor import (
    _is_date_like,
    extract_cells,
    has_cell_grid,
)


def _grid_ws():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Name")
    for col, day in enumerate(range(1, 4), start=2):
        ws.cell(row=1, column=col, value=dt.date(2026, 6, day))
    ws.cell(row=2, column=1, value="John")
    ws.cell(row=2, column=2, value="D")
    ws.cell(row=2, column=3, value="N")
    ws.cell(row=2, column=4, value="OFF")
    ws.cell(row=3, column=1, value="Jane")
    ws.cell(row=3, column=2, value="OFF")
    ws.cell(row=3, column=4, value="D")
    return ws


def test_is_date_like():
    assert _is_date_like(dt.date(2026, 6, 1))
    assert _is_date_like(15)
    assert _is_date_like("Mon")
    assert _is_date_like("Tuesday")
    assert not _is_date_like("OFF")
    assert not _is_date_like(True)
    assert not _is_date_like(99)


def test_has_cell_grid():
    assert has_cell_grid(_grid_ws())


def test_extract_basic_grid():
    res = extract_cells(_grid_ws())
    assert res.date_axis == ["2026-06-01", "2026-06-02", "2026-06-03"]
    assert [p.name for p in res.people] == ["John", "Jane"]

    john = {s.date: s.raw_code for s in res.people[0].shifts}
    assert john == {
        "2026-06-01": "D",
        "2026-06-02": "N",
        "2026-06-03": "OFF",
    }

    # Jane has a blank middle day, which should be skipped (not emitted as empty).
    jane = {s.date: s.raw_code for s in res.people[1].shifts}
    assert jane == {"2026-06-01": "OFF", "2026-06-03": "D"}


def test_header_and_name_overrides():
    wb = openpyxl.Workbook()
    ws = wb.active
    # Junk first row, real header on row 2, names in column B.
    ws.cell(row=1, column=1, value="ignore me")
    ws.cell(row=2, column=2, value="Worker")
    ws.cell(row=2, column=3, value="Mon")
    ws.cell(row=2, column=4, value="Tue")
    ws.cell(row=3, column=2, value="Sam")
    ws.cell(row=3, column=3, value="D")
    ws.cell(row=3, column=4, value="N")

    res = extract_cells(ws, header_row=2, name_col="B")
    assert [p.name for p in res.people] == ["Sam"]
    codes = {s.date: s.raw_code for s in res.people[0].shifts}
    assert codes == {"Mon": "D", "Tue": "N"}
