"""Layout quirks of the real workbooks that the parser has to survive.

Everything here is modelled on the shipped files: a tab whose name is not the
period title, whole weekend columns shaded pale green, and day boxes drawn as a
single fill spanning two rows. Each one of these used to make the parser produce
confidently wrong output.
"""

import openpyxl
from openpyxl.styles import Color, PatternFill

from schedule_extractor.definitions import decode, fill_key, has_solid_fill
from schedule_extractor.roster_extractor import extract_roster

WEEKEND = PatternFill("solid", fgColor="FFCCFFCC")   # weekend column shading
BOX = PatternFill("solid", fgColor="FFCCC0DA")       # a person's day box
BAR = PatternFill("solid", fgColor="FFFFFF00")       # a split's center bar
APPROVED = PatternFill("solid", fgColor="FF00B050")  # approved-vacation green


def _sheet(tab_title, row1_title=None, start_day=13, days=14, first_dow="S"):
    """A roster sheet whose dates start on Sunday Sept 13, 2026."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab_title
    if row1_title:
        ws.cell(1, 2, row1_title)
    ws.cell(2, 1, "DATE")
    ws.cell(3, 1, "DAY")
    letters = ["S", "M", "T", "W", "Th", "F", "S"]
    offset = letters.index(first_dow) if first_dow in letters else 0
    for i in range(days):
        ws.cell(2, 2 + i, start_day + i)
        ws.cell(3, 2 + i, letters[(offset + i) % 7])
    return ws


def _person(ws, row, name, contacts=("C: 555-1212", "P: TXT")):
    ws.cell(row, 1, name)
    for i, c in enumerate(contacts):
        ws.cell(row + 1 + i, 1, c)


# --- the date axis -------------------------------------------------------


def test_period_title_comes_from_row_1_not_the_tab_name():
    """A "Working…" tab prefix must not be read as the month.

    The shipped tab is named "WorkingSept13 - Oct 10,  (4)": scanning for the
    first three letters finds "Wor", which is not a month, and the whole
    schedule silently lands in January.
    """
    ws = _sheet("WorkingSept13 - Oct 10,  (4)", "September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "BC")

    res = extract_roster(ws)
    assert res["date_range"]["start"] == "2026-09-13"
    assert res["title"] == "September 13 - October 10, 26"
    assert res["warnings"] == []


def test_month_name_is_matched_anywhere_in_the_title():
    ws = _sheet("Draft Sept 13 - Oct 10, 26")     # "Dra" is not a month
    _person(ws, 4, "SMITH")
    assert extract_roster(ws)["date_range"]["start"] == "2026-09-13"


def test_day_row_fills_in_a_year_the_title_never_stated():
    """Sept 13 is a Sunday in 2026 but a Saturday in 2025.

    A bare trailing "10" ("Sept 13 - Oct 10") is a day of the month, not 2010.
    """
    ws = _sheet("Sept 13 - Oct 10")  # no year anywhere
    _person(ws, 4, "SMITH")

    res = extract_roster(ws, default_year=2025)
    assert res["date_range"]["start"] == "2026-09-13"
    assert any("matches 2026" in w for w in res["warnings"])


def test_a_year_stated_in_the_title_is_not_overridden():
    """The owner's own year stands; a weekday clash is reported, not corrected."""
    ws = _sheet("September 13 - October 10, 25", first_dow="S")
    _person(ws, 4, "SMITH")

    res = extract_roster(ws)
    assert res["date_range"]["start"] == "2025-09-13"
    assert any("disagrees with the DAY row" in w for w in res["warnings"])


def test_unreadable_month_warns_instead_of_silently_using_january():
    ws = _sheet("Working copy (4)")
    _person(ws, 4, "SMITH")

    res = extract_roster(ws)
    assert any("could not read a month" in w for w in res["warnings"])


def test_incoherent_day_row_is_reported():
    """A DAY row no year can satisfy is flagged rather than quietly accepted."""
    ws = _sheet("Sept 13 - Oct 10")
    for col in range(2, 16):
        ws.cell(3, col, "M")                      # every column claims Monday
    _person(ws, 4, "SMITH")

    res = extract_roster(ws)
    assert any("disagrees with the DAY row" in w for w in res["warnings"])


# --- weekend shading vs. real colour -------------------------------------


def test_weekend_shading_does_not_approve_a_vacation():
    """Sat/Sun columns are shaded green for everyone; that is not an approval.

    Reading the shading as approval splits one vacation block into "approved" on
    its weekend days and "not approved" on its weekdays.
    """
    ws = _sheet("September 13 - October 10, 26")
    for row in range(4, 16):                      # three people, 3 rows each
        for col in (2, 8, 9):                     # Sun 13, Sat 19, Sun 20
            ws.cell(row, col).fill = WEEKEND
    _person(ws, 4, "SMITH")
    _person(ws, 7, "JONES")
    _person(ws, 10, "PATEL")
    for col in range(7, 11):                      # Fri 18 .. Sun 21, one block
        ws.cell(4, col, "V")

    vacation = {s["date"]: s for s in extract_roster(ws)["people"][0]["shifts"]}
    assert [s["approved"] for s in vacation.values()] == [False] * 4


def test_a_green_that_is_not_the_column_shading_still_approves():
    ws = _sheet("September 13 - October 10, 26")
    for row in range(4, 16):
        for col in (2, 8, 9):
            ws.cell(row, col).fill = WEEKEND
    _person(ws, 4, "SMITH")
    _person(ws, 7, "JONES")
    _person(ws, 10, "PATEL")
    ws.cell(4, 7, "V")                            # Fri 18, plain
    approved = ws.cell(4, 8, "V")                 # Sat 19, deliberately green
    approved.fill = APPROVED

    by_date = {s["date"]: s for s in extract_roster(ws)["people"][0]["shifts"]}
    assert by_date["2026-09-18"]["approved"] is False
    assert by_date["2026-09-19"]["approved"] is True


# --- day boxes vs. split center bars -------------------------------------


def test_two_row_day_box_is_not_a_split():
    """The day box is two rows tall and filled in one colour.

    Treating "the middle row has a fill" as a split flags every worked day in the
    sheet, and halves each clinic day to a morning.
    """
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "CNV").fill = BOX
    ws.cell(5, 2).fill = BOX                      # same box, no code

    shift = extract_roster(ws)["people"][0]["shifts"][0]
    assert "split_day" not in shift
    assert (shift["start"], shift["end"]) == ("08:00", "17:00")


def test_second_line_of_a_box_is_a_note_not_a_midshift():
    """"APP" over "C" is one label ("APP C"), not a day plus an afternoon."""
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "APP").fill = BOX
    ws.cell(5, 2, "C").fill = BOX

    person = extract_roster(ws)["people"][0]
    assert [s["shift_type"] for s in person["shifts"]] == ["day"]
    assert person["shifts"][0]["meaning"] == "LJ OBGYN APP C (Triage/PP)"
    assert (person["shifts"][0]["start"], person["shifts"][0]["end"]) == ("06:30", "19:00")
    assert {"date": "2026-09-13", "text": "C"} in person["notes"]


def test_center_bar_in_a_different_colour_is_still_a_split():
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "MOS").fill = BOX
    ws.cell(5, 2, "CNV").fill = BAR               # a real center bar

    by = {(s["code"], s["shift_type"]): s for s in extract_roster(ws)["people"][0]["shifts"]}
    morning, afternoon = by[("MOS", "day")], by[("CNV", "midshift")]
    assert morning["split_day"] is True
    assert (morning["start"], morning["end"]) == ("08:00", "12:00")
    assert (afternoon["start"], afternoon["end"]) == ("13:00", "17:00")


# --- codes ---------------------------------------------------------------


def test_cnv_is_convoy_and_counts_as_a_clinic():
    """The workbooks spell Convoy "CNV"; only "CV" was defined."""
    assert decode("CNV") == {"category": "location", "meaning": "Convoy"}
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "CNV")
    shift = extract_roster(ws)["people"][0]["shifts"][0]
    assert (shift["start"], shift["end"]) == ("08:00", "17:00")


def test_four_character_codes_are_reachable():
    """NTAS and BDay are defined but the 1-3 char cap made them unreachable."""
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "BDay")
    ws.cell(6, 3, "NTAS")

    shifts = {s["code"]: s for s in extract_roster(ws)["people"][0]["shifts"]}
    assert shifts["BDay"]["category"] == "status"
    assert shifts["NTAS"]["category"] == "location"
    assert shifts["NTAS"]["shift_type"] == "night"


def test_codes_are_decoded_case_insensitively():
    assert decode("bday")["meaning"] == "Birthday request (off)"
    assert decode("cnv")["meaning"] == "Convoy"


def test_counted_code_keeps_the_shift_and_the_counter():
    """"BC6" is an orientation shift, not free text — it used to be dropped."""
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(6, 2, "BC6")

    person = extract_roster(ws)["people"][0]
    assert person["shifts"][0]["code"] == "BC"
    assert person["shifts"][0]["shift_type"] == "night"
    assert {"date": "2026-09-13", "text": "BC6"} in person["notes"]


def test_a_counter_never_turns_free_text_into_a_shift():
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "XY7")

    person = extract_roster(ws)["people"][0]
    assert person["shifts"] == []
    assert person["notes"] == [{"date": "2026-09-13", "text": "XY7"}]


# --- fills ---------------------------------------------------------------


def test_theme_coloured_fills_are_visible():
    """Only fgColor.rgb was read, so theme fills looked like no fill at all."""
    wb = openpyxl.Workbook()
    ws = wb.active
    themed = ws.cell(1, 1)
    themed.fill = PatternFill(patternType="solid", fgColor=Color(theme=9, tint=0.0))
    white = ws.cell(1, 2)
    white.fill = PatternFill(patternType="solid", fgColor=Color(theme=0, tint=0.0))

    assert has_solid_fill(themed) is True
    assert has_solid_fill(white) is False          # theme 0 is the white background
    assert fill_key(themed) == ("theme", 9, 0.0)
    assert fill_key(themed) != fill_key(white)
    assert fill_key(ws.cell(1, 3)) is None         # unfilled


def test_cv_and_cnv_are_matched_as_one_clinic():
    """A roster written "CV" must still match a schedule that says "CNV"."""
    from schedule_extractor.definitions import canonical_code
    from server.coverage import _qualified_for

    assert canonical_code("cv") == "CNV"
    assert canonical_code("CNV") == "CNV"
    assert canonical_code("VLJ") == "VLJ"

    meta = {"clinics": {canonical_code("CV")}}
    assert _qualified_for("CNV", {"codes": set()}, meta) is True
    assert _qualified_for("CV", {"codes": {"CNV"}}, None) is True


def test_leap_day_is_resolved_to_a_year_that_has_one():
    """Feb 29 against a non-leap year used to raise out of the middle of a parse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "February 27 - March 2"                # no year stated
    ws.cell(2, 1, "DATE")
    ws.cell(3, 1, "DAY")
    for i, (day, dow) in enumerate([(27, "S"), (28, "M"), (29, "T"), (1, "W"), (2, "Th")]):
        ws.cell(2, 2 + i, day)
        ws.cell(3, 2 + i, dow)
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "BC")

    res = extract_roster(ws, default_year=2027)       # 2027 has no Feb 29
    assert res["date_range"] == {"start": "2028-02-27", "end": "2028-03-02"}
    assert res["people"][0]["shifts"][0]["code"] == "BC"


def test_a_month_that_cannot_hold_the_dates_warns_and_still_parses():
    """No year has a February 30th — say so rather than raising."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "February 27 - February 30, 26"
    ws.cell(2, 1, "DATE")
    for i, day in enumerate([27, 28, 29, 30]):
        ws.cell(2, 2 + i, day)
    _person(ws, 4, "SMITH")
    ws.cell(4, 2, "BC")

    res = extract_roster(ws)
    assert any("cannot hold the day numbers" in w for w in res["warnings"])
    assert res["people"][0]["shifts"][0]["code"] == "BC"   # data still comes through


# --- codes the owner confirmed ------------------------------------------


def test_owner_confirmed_codes():
    """E, MC, NRP and JD, as defined by the schedule owner.

    MC is a location and follows the "anything that isn't BC or HC is a clinic"
    rule, so it gets the standard clinic day. E/NRP/JD are commitments away from
    the floor: no clock window, and the coverage engine treats them as off.
    """
    ws = _sheet("September 13 - October 10, 26")
    _person(ws, 4, "SMITH")
    for col, code in [(2, "E"), (3, "MC"), (4, "NRP"), (5, "JD")]:
        ws.cell(4, col, code)

    by = {s["code"]: s for s in extract_roster(ws)["people"][0]["shifts"]}
    assert by["MC"]["category"] == "location"
    assert by["MC"]["meaning"] == "Mid City"
    assert (by["MC"]["start"], by["MC"]["end"]) == ("08:00", "17:00")
    for code, meaning in [("E", "Education"), ("JD", "Jury duty")]:
        assert by[code]["category"] == "status"
        assert by[code]["meaning"] == meaning
        assert by[code]["start"] is None
    assert by["NRP"]["meaning"].startswith("NRP course")


def test_a_commitment_elsewhere_is_not_offered_as_a_cover():
    """Someone at jury duty or a course must not be proposed to cover a shift."""
    from server.coverage import _OFF_REASON, _OFF_STATUS

    for code in ("E", "NRP", "JD"):
        assert code in _OFF_STATUS
        assert code in _OFF_REASON
