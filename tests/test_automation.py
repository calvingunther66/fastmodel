import io

import openpyxl

from server import mcp
from server.apitokens import ApiTokenStore
from server.automation import Automation
from server.store import ScheduleStore


def _xlsx_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "June 21 - July 18, 26"
    ws.cell(2, 1, "DATE")
    ws.cell(2, 2, 21)
    ws.cell(2, 3, 22)
    ws.cell(4, 1, "DOE")
    ws.cell(4, 2, "BC")
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_api_tokens(tmp_path):
    s = ApiTokenStore(tmp_path)
    rec, tok = s.create("agent", ["automate", "bogus"])
    assert rec["capabilities"] == ["automate"]          # invalid cap filtered out
    assert tok.startswith("sk_sched_")

    p = s.verify(tok)
    assert p and p["is_token"] and p["capabilities"] == ["automate"]
    assert s.verify("sk_sched_nope") is None
    assert s.verify("not-even-prefixed") is None
    assert all("token_hash" not in t for t in s.list())  # hash never exposed

    s.revoke(rec["id"])
    assert s.verify(tok) is None


def test_automation_ingest_latest(tmp_path):
    store = ScheduleStore(tmp_path)
    inbox = tmp_path / "inbox"
    inbox.mkdir()
    auto = Automation(store, inbox=inbox)

    assert auto.ingest_latest()["status"] == "empty"

    (inbox / "schedule.xlsx").write_bytes(_xlsx_bytes())
    added = auto.ingest_latest()
    assert added["status"] == "added"
    assert added["people"] >= 1
    assert added["period"] == "2026-06-21..2026-06-22"

    # Same content again -> idempotent no-op.
    assert auto.ingest_latest()["status"] == "unchanged"

    st = auto.status()
    assert st["spreadsheets"] == 1
    assert "2026-06-21..2026-06-22" in st["periods_ingested"]


class _Auto:
    def list_spreadsheets(self): return [{"name": "x.xlsx"}]
    def ingest_latest(self, **k): return {"status": "added"}
    def status(self): return {"inbox": "/data/inbox", "periods_ingested": []}


class _Store:
    def get_schedule(self):
        return {"parsed_sheet": "June", "date_range": {"start": "a"}, "people": [{"name": "A"}]}


def test_mcp_handle():
    a, s = _Auto(), _Store()
    init = mcp.handle({"jsonrpc": "2.0", "id": 1, "method": "initialize"}, a, s)
    assert init["result"]["serverInfo"]["name"] == "fastmodel-scheduler"

    tools = mcp.handle({"id": 2, "method": "tools/list"}, a, s)["result"]["tools"]
    assert {t["name"] for t in tools} == {"list_spreadsheets", "ingest_latest", "schedule_status"}

    call = mcp.handle({"id": 3, "method": "tools/call",
                       "params": {"name": "schedule_status"}}, a, s)
    assert "active_sheet" in call["result"]["content"][0]["text"]

    # notifications get no response; unknown methods are JSON-RPC errors
    assert mcp.handle({"method": "notifications/initialized"}, a, s) is None
    err = mcp.handle({"id": 9, "method": "nope"}, a, s)
    assert err["error"]["code"] == -32601
